# app.py
import os
import pandas as pd

BASE_DIR=os.path.dirname(__file__)
EXCEL_PATH = os.path.join(BASE_DIR, "bus_hersteller_zuordnung.xlsx")

mapping_df = pd.read_excel(EXCEL_PATH, engine="openpyxl")

from pathlib import Path
import numpy as np
import pandas as pd
import streamlit as st
import json
import plotly.express as px
import plotly.io as pio
from typing import Tuple, Dict, Any, Optional
from io import BytesIO
import calendar
from openpyxl.styles import PatternFill, Alignment
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
from matplotlib.colors import ListedColormap
import pandas as pd
from io import BytesIO
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.chart import LineChart, Reference
from Funktionen import (
    load_data,
    to_excel_raw,
    DISCRETE_SCHEMAS,
    CONTINUOUS_SCHEMAS
)



FEHLERKATEGORIEN: Dict[str, Optional[str]] = {
    "Hess": "Hess",
    "Neufahrn": "MAN",
    "Abdeckung": "Innenraum",
    "Abgasanlage": "Abgasanlage",
    "Abgasrohr": "Abgasanlage",
    "ABS": "Bremse",
    "Absperrband": "Anh√§nger",
    "Absperrung": "Anh√§nger",
    "Achse": "Achsen",
    "AdBlue": "Motor",
    "AGR": "Abgasanlage",
    "AGR Ventil": "Abgasanlage",
    "AHK": "Anh√§nger",
    "Akku": "Batterie",
    "Amaturen": "Fahrerarbeitsplatz",
    "Anf. Sperre": "Bremse",
    "Anlasser": "Motor",
    "Antrieb": "Antrieb",
    "Anzeige": "Innenraum",
    "Armaturenbrett": "Fahrerarbeitsplatz",
    "ASR": "Bremse",
    "Assistenzsysteme": "Elektrik",
    "Aufbau": "Luftfederung",
    "Auspuff": "Abgasanlage",
    "Batterie": "Batterie",
    "Beklebung": "SWM",
    "Beleuchtung": "Lichtanlage",
    "Blech": "Karrosserie",
    "Bodenbelag": "Innenraum",
    "Brand": "Brand",
    "Brandgeruch": "Brand",
    "Brandmeldeanlage": "Brand",
    "Bremse": "Bremse",
    "Bremspedal": "Bremse",
    "CAN": "Elektrik",
    "Copilot": "INIT",
    "CRT": "Abgasanlage",
    "Dachklappe": "Innenraum",
    "Dachluke": "Innenraum",
    "Dachverkleidung": "Innenraum",
    "Defekt": "Defekt",
    "Deichsel": "Anh√§nger",
    "Differential": "Antrieb",
    "Di-Mi": None,
    "Display": "Innenraum",
    "DPF": "Abgasanlage",
    "Drehkranz": "Drehkranz",
    "EBS": "Bremse",
    "ECAS": "Luftfederung",
    "EDC": "Motor",
    "EKS": None,
    "Elektrik": "Elektrik",
    "ESP": "Bremse",
    "E-Stecker": "Elektrik",
    "Fahrerfenster": "Fahrerarbeitsplatz",
    "Fahrersitz": "Fahrerarbeitsplatz",
    "Fahrert√ºre": "Fahrerarbeitsplatz",
    "Faltenbalg": "Drehkranz",
    "Fenster": "Fenster",
    "FFR": None,
    "Fingerschutz": "T√ºre",
    "Flexrohr": "Abgasanlage",
    "Gebl√§se": "Heizung",
    "Generator": "Elektrik",
    "Ger√§usch": None,
    "Getriebe": "Antrieb",
    "Haltestange": "Innenraum",
    "Heizung": "Heizung",
    "Himmel": "Innenraum",
    "HLK": "Heizung",
    "Hupe": "Fahrerarbeitsplatz",
    "INET": "INIT",
    "INIT": "INIT",
    "K 1": "Reparatur",
    "K 4": "Reparatur",
    "K 6": "Reparatur",
    "K1": "Reparatur",
    "K4": "Reparatur",
    "K6": "Reparatur",
    "Kabel": "Elektrik",
    "Kabelschaden": "Elektrik",
    "Kamera": "Video",
    "Kardanwelle": "Antrieb",
    "Keilriemen": "Motor",
    "Kennung": "Anh√§nger",
    "Klappe": "Klapprampe",
    "Klapprampe": "Klapprampe",
    "Knickschutz": "Drehkranz",
    "Kompressor": "Druckluftanlage",
    "Kraftstoff": "Kraftstoff",
    "K√ºhler": "K√ºhler",
    "K√ºhlerklappe": "Karrosserie",
    "Ladeluftk√ºhler": "K√ºhler",
    "Laden": "Elektrik",
    "Ladung": "SOC",
    "Leiste": "Innenraum",
    "Leistung": "Antrieb",
    "Leitung": "Elektrik",
    "Lenkrad": "Fahrerarbeitsplatz",
    "LENKUNG": "Lenkung",
    "Licht": "Lichtanlage",
    "Lima": "Klimaanlage",
    "LiMa Kissen": "Klimaanlage",
    "Luftanlage": "Druckluftanlage",
    "Luftfederung": "Luftfederung",
    "Luftpresser": "Druckluftanlage",
    "L√ºftung": "Klimaanlage",
    "Luftverlust": "Druckluftanlage",
    "Mikrofon": "Fahrerarbeitsplatz",
    "Monitor": "Innenraum",
    "Motor": "Motor",
    "Motorabdeckung": "Motor",
    "Motork√ºhlung": "Motor",
    "MR": "Motor",
    "NR": "Reparatur",
    "NR Bremse": "Bremse",
    "NR Sa/So": "Reparatur",
    "√ñlverlust": "Motor",
    "Polster": "Sitze",
    "Reifen": "Reifen",
    "Sa/Mo": "Reparatur",
    "Sa/So": "Reparatur",
    "Schmierung": "Antrieb",
    "Seitendeckel": "Karrosserie",
    "Sitz": "Sitze",
    "S-Klappe": "Karrosserie",
    "Spannung": "Elektrik",
    "Spiegel": "Spiegel",
    "Spur": "Achsen",
    "Spurstange": "Fahrwerk",
    "Startanlage": "Motor",
    "Stecker": "Elektrik",
    "St√∂rung": "Defekt",
    "Sto√üd√§mpfer": "Fahrwerk",
    "SWF": None,
    "System": "Elektrik",
    "TA": "Achsen",
    "Tachogeber": "Fahrerarbeitsplatz",
    "Tank": "Kraftstoff",
    "Tankanz.": "Kraftstoff",
    "Traverse": "Motor",
    "T√ºr": "T√ºre",
    "Turbo": "Motor",
    "T√ºre": "T√ºre",
    "Umlenkh.": "Lenkung",
    "Umlenkhebel": "Lenkung",
    "Unterbodens": "Unterbodenschutz",
    "Unterbodens.": "Unterbodenschutz",
    "Unterschutz": "Unterbodenschutz",
    "VA": "Achsen",
    "Verkleidung": "Innenraum",
    "Video": "SWM",
    "Voith": "Antrieb",
    "Voith Monitor": "Antrieb",
    "Wapu": "K√ºhlwasser",
    "Wartung Sa/So": "Wartung",
    "Wasser": "K√ºhlwasser",
    "Wassereinbruch": "K√ºhlwasser",
    "Wasserverl": "K√ºhlwasser",
    "Wasserverl.": "K√ºhlwasser",
    "Wasserverlust": "K√ºhlwasser",
    "Webasto": "Heizung",
    "Wischer": "Scheibenwischer",
    "Zahnkranz": "Motor",
    "Zentral Schmierung": "Motor",
    "Z-Fzg": "Anh√§nger",
    "Zielanzeige": "Zielanzeige",
    "Zielschild": "Zielanzeige",
    "Z-Schmier.": "Motor",
    "Zusatzh.": "Heizung",

    # Kategorie 3
    "Klima": "Klimaanlage",
    "Nachunt.": "SWM-Abnahme",
    "Nicht gekoppelt": "Anh√§nger",
    "NR Wartung": "Wartung",
    "Sp": "SWM-Abnahme",
    "Wartung": "Wartung",
    "Wartung MR": "Wartung",

    # Kategorie 4
    "AU": "SWM-Abnahme",
    "HU": "SWM-Abnahme",
    "Tacho": "SWM-Abnahme",
    "T√úV": "SWM-Abnahme",

    # Kategorie 5
    "Corona": None,
    "HV": "Anh√§nger",
    "Kein Z-Fzg": "Anh√§nger",
    "Quarant√§ne": None,
    "Schneeketten": None,

    # Kategorie 6
    "Reiniger": "SWM",
    "Reinigung": "SWM",
    "Reinigung Sa": "SWM",
    "Reklame": "SWM",
    "Sa Reklame": "SWM",

    # Kategorie 7
    "Lack": "Unfall",
    "Lackierer": "Unfall",
    "Schaden": "Unfall",
    "Scheibe": "Fenster",
    "Speng": "Unfall",
    "Spengler": "Unfall",
    "Unfall": "Unfall",

    # Kategorie 8
    "Aufr√ºstung": "SWM",
    "AW": "SWM",
    "BBH West": "SWM",
    "Bhf Ost": "SWM",
    "Bhf West": "SWM",
    "Entwerter": "SWM",
    "FA": "SWM",
    "Fahrschule": "SWM",
    "Film": "SWM",
    "Foto": "SWM",
    "Funk": "SWM",
    "Piktogramme": "SWM",
    "Pixida": "SWM",
    "Schulung": "SWM",
    "Sonderf.": "SWM",
    "Sonderfahrt": "SWM",
    "Stelzer": "SWM",
    "Telematik": "SWM",
    "Um-bau": "SWM",
    "Vermietet": "SWM",

    # Kategorie 9
    "AOB": "Fremdfirma",
    "B u. M": "Fremdfirma",
    "Beissbarth": "Fremdfirma",
    "Edelwei√ü": "Fremdfirma",
    "Fa Bode": "Fremdfirma",
    "Fa Knorr": "Fremdfirma",
    "Fa Lawo": "Fremdfirma",
    "Fa Mayo": "Fremdfirma",
    "Gr√§felfing": "Fremdfirma",
    "Gruber": "Fremdfirma",
    "HJS": "Fremdfirma",
    "H√ºttinger": "Fremdfirma",
    "Kienzle": "Fremdfirma",
    "Kronberger": "Fremdfirma",
    "KWS": "Fremdfirma",
    "ML": "Fremdfirma",
    "√ñPNV": "Fremdfirma",
    "Ribas": "Fremdfirma",
    "Romaldini": "Fremdfirma",
    "Sinos": "Fremdfirma",
    "Spusi": "Fremdfirma",
    "TrailerServ": "Fremdfirma",
    "Tu.T": "Fremdfirma",
}



AUTO_FEHLERKATEGORIEN = {
    # 1 ‚Äì Hess / MAN
    "Hess":         (1, "Hess"),
    "Neufahrn":     (1, "MAN"),

    # 2 ‚Äì alle sonstigen Kategorien
    "Abdeckung":            (2, "Innenraum"),
    "Abgasanlage":          (2, "Abgasanlage"),
    "Abgasrohr":            (2, "Abgasanlage"),
    "ABS":                  (2, "Bremse"),
    "Absperrband":          (2, "Anh√§nger"),
    "Absperrung":           (2, "Anh√§nger"),
    "Achse":                (2, "Achsen"),
    "AdBlue":               (2, "Motor"),
    "AGR":                  (2, "Abgasanlage"),
    "AGR Ventil":           (2, "Abgasanlage"),
    "AHK":                  (2, "Anh√§nger"),
    "Akku":                 (2, "Batterie"),
    "Amaturen":             (2, "Fahrerarbeitsplatz"),
    "Anf. Sperre":          (2, "Bremse"),
    "Anlasser":             (2, "Motor"),
    "Antrieb":              (2, "Antrieb"),
    "Anzeige":              (2, "Innenraum"),
    "Armaturenbrett":       (2, "Fahrerarbeitsplatz"),
    "ASR":                  (2, "Bremse"),
    "Assistenzsysteme":     (2, "Elektrik"),
    "Aufbau":               (2, "Luftfederung"),
    "Auspuff":              (2, "Abgasanlage"),
    "Batterie":             (2, "Batterie"),
    "Beklebung":            (2, "SWM"),
    "Beleuchtung":          (2, "Lichtanlage"),
    "Blech":                (2, "Karrosserie"),
    "Bodenbelag":           (2, "Innenraum"),
    "Brand":                (2, "Brand"),
    "Brandgeruch":          (2, "Brand"),
    "Brandmeldeanlage":     (2, "Brand"),
    "Bremse":               (2, "Bremse"),
    "Bremspedal":           (2, "Bremse"),
    "CAN":                  (2, "Elektrik"),
    "Copilot":              (2, "INIT"),
    "CRT":                  (2, "Abgasanlage"),
    "Dachklappe":           (2, "Innenraum"),
    "Dachluke":             (2, "Innenraum"),
    "Dachverkleidung":      (2, "Innenraum"),
    "Defekt":               (2, "Defekt"),
    "Deichsel":             (2, "Anh√§nger"),
    "Differential":         (2, "Antrieb"),
    "Di-Mi":                (2, "Sonstiges"),
    "Display":              (2, "Innenraum"),
    "DPF":                  (2, "Abgasanlage"),
    "Drehkranz":            (2, "Drehkranz"),
    "EBS":                  (2, "Bremse"),
    "ECAS":                 (2, "Luftfederung"),
    "EDC":                  (2, "Motor"),
    "EKS":                  (2, "Sonstiges"),
    "Elektrik":             (2, "Elektrik"),
    "ESP":                  (2, "Bremse"),
    "E-Stecker":            (2, "Elektrik"),
    "Fahrerfenster":        (2, "Fahrerarbeitsplatz"),
    "Fahrersitz":           (2, "Fahrerarbeitsplatz"),
    "Fahrert√ºre":           (2, "Fahrerarbeitsplatz"),
    "Faltenbalg":           (2, "Drehkranz"),
    "Fenster":              (2, "Fenster"),
    "FFR":                  (2, "Sonstiges"),
    "Fingerschutz":         (2, "T√ºre"),
    "Flexrohr":             (2, "Abgasanlage"),
    "Gebl√§se":              (2, "Heizung"),
    "Generator":            (2, "Elektrik"),
    "Ger√§usch":             (2, "Sonstiges"),
    "Getriebe":             (2, "Antrieb"),
    "Haltestange":          (2, "Innenraum"),
    "Heizung":              (2, "Heizung"),
    "Himmel":               (2, "Innenraum"),
    "HLK":                  (2, "Heizung"),
    "Hupe":                 (2, "Fahrerarbeitsplatz"),
    "INET":                 (2, "INIT"),
    "INIT":                 (2, "INIT"),
    "K 1":                  (2, "Reparatur"),
    "K 4":                  (2, "Reparatur"),
    "K 6":                  (2, "Reparatur"),
    "K1":                   (2, "Reparatur"),
    "K4":                   (2, "Reparatur"),
    "K6":                   (2, "Reparatur"),
    "Kabel":                (2, "Elektrik"),
    "Kabelschaden":         (2, "Elektrik"),
    "Kamera":               (2, "Video"),
    "Kardanwelle":          (2, "Antrieb"),
    "Keilriemen":           (2, "Motor"),
    "Kennung":              (2, "Anh√§nger"),
    "Klappe":               (2, "Klapprampe"),
    "Klapprampe":           (2, "Klapprampe"),
    "Knickschutz":          (2, "Drehkranz"),
    "Kompressor":           (2, "Druckluftanlage"),
    "Kraftstoff":           (2, "Kraftstoff"),
    "K√ºhler":               (2, "K√ºhler"),
    "K√ºhlerklappe":         (2, "Karrosserie"),
    "Ladeluftk√ºhler":       (2, "K√ºhler"),
    "Laden":                (2, "Elektrik"),
    "Ladung":               (2, "SOC"),
    "Leiste":               (2, "Innenraum"),
    "Leistung":             (2, "Antrieb"),
    "Leitung":              (2, "Elektrik"),
    "Lenkrad":              (2, "Fahrerarbeitsplatz"),
    "LENKUNG":              (2, "Lenkung"),
    "Licht":                (2, "Lichtanlage"),
    "Lima":                 (2, "Klimaanlage"),
    "LiMa Kissen":          (2, "Klimaanlage"),
    "Luftanlage":           (2, "Druckluftanlage"),
    "Luftfederung":         (2, "Luftfederung"),
    "Luftpresser":          (2, "Druckluftanlage"),
    "L√ºftung":              (2, "Klimaanlage"),
    "Luftverlust":          (2, "Druckluftanlage"),
    "Mikrofon":             (2, "Fahrerarbeitsplatz"),
    "Monitor":              (2, "Innenraum"),
    "Motor":                (2, "Motor"),
    "Motorabdeckung":       (2, "Motor"),
    "Motork√ºhlung":         (2, "Motor"),
    "MR":                   (2, "Motor"),
    "NR":                   (2, "Reparatur"),
    "NR Bremse":            (2, "Bremse"),
    "NR Sa/So":             (2, "Reparatur"),
    "√ñlverlust":            (2, "Motor"),
    "Polster":              (2, "Sitze"),
    "Reifen":               (2, "Reifen"),
    "Sa/Mo":                (2, "Reparatur"),
    "Sa/So":                (2, "Reparatur"),
    "Schmierung":           (2, "Antrieb"),
    "Seitendeckel":         (2, "Karrosserie"),
    "Sitz":                 (2, "Sitze"),
    "S-Klappe":             (2, "Karrosserie"),  # ‚ÄûS-Klappe‚Äú
    "Spannung":             (2, "Elektrik"),
    "Spiegel":              (2, "Spiegel"),
    "Spur":                 (2, "Achsen"),
    "Spurstange":           (2, "Fahrwerk"),
    "Startanlage":          (2, "Motor"),
    "Stecker":              (2, "Elektrik"),
    "St√∂rung":              (2, "Defekt"),
    "Sto√üd√§mpfer":          (2, "Fahrwerk"),
    "SWF":                  (2, "Sonstiges"),
    "System":               (2, "Elektrik"),
    "TA":                   (2, "Achsen"),
    "Tachogeber":           (2, "Fahrerarbeitsplatz"),
    "Tank":                 (2, "Kraftstoff"),
    "Tankanz.":             (2, "Kraftstoff"),
    "Traverse":             (2, "Motor"),
    "T√ºr":                  (2, "T√ºre"),
    "Turbo":                (2, "Motor"),
    "T√ºre":                 (2, "T√ºre"),
    "Umlenkh.":             (2, "Lenkung"),
    "Umlenkhebel":          (2, "Lenkung"),
    "Unterbodens":          (2, "Unterbodenschutz"),
    "Unterbodens.":         (2, "Unterbodenschutz"),
    "Unterschutz":          (2, "Unterbodenschutz"),
    "VA":                   (2, "Achsen"),
    "Verkleidung":          (2, "Innenraum"),
    "Video":                (2, "SWM"),
    "Voith":                (2, "Antrieb"),
    "Voith Monitor":        (2, "Antrieb"),
    "Wapu":                 (2, "K√ºhlwasser"),
    "Wartung Sa/So":        (2, "Wartung"),
    "Wasser":               (2, "K√ºhlwasser"),
    "Wassereinbruch":       (2, "K√ºhlwasser"),
    "Wasserverl":           (2, "K√ºhlwasser"),
    "Wasserverl.":          (2, "K√ºhlwasser"),
    "Wasserverlust":        (2, "K√ºhlwasser"),
    "Webasto":              (2, "Heizung"),
    "Wischer":              (2, "Scheibenwischer"),
    "Zahnkranz":            (2, "Motor"),
    "Zentral Schmierung":   (2, "Motor"),
    "Z-Fzg":                (2, "Anh√§nger"),
    "Zielanzeige":          (2, "Zielanzeige"),
    "Zielschild":           (2, "Zielanzeige"),
    "Z-Schmier.":           (2, "Motor"),      # ‚ÄûZ-Schmier.‚Äú
    "Zusatzh.":             (2, "Heizung"),

    # 3 ‚Äì Wartung & SWM-Abnahme
    "Klima":                (3, "Klimaanlage"),
    "Nachunt.":             (3, "SWM-Abnahme"),
    "Nicht gekoppelt":      (3, "Anh√§nger"),
    "NR Wartung":           (3, "Wartung"),
    "Sp":                   (3, "SWM-Abnahme"),
    "Wartung":              (3, "Wartung"),
    "Wartung MR":           (3, "Wartung"),

    # 4 ‚Äì SWM-Abnahme
    "AU":                   (4, "SWM-Abnahme"),
    "HU":                   (4, "SWM-Abnahme"),
    "Tacho":                (4, "SWM-Abnahme"),
    "T√úV":                  (4, "SWM-Abnahme"),

    # 5 ‚Äì Quarant√§ne, Corona, HV, Anh√§nger-Ausnahmen
    "Corona":               (5, "Sonstiges"),
    "HV":                   (5, "Anh√§nger"),
    "Kein Z-Fzg":           (5, "Anh√§nger"),
    "Quarant√§ne":           (5, "Sonstiges"),

    # 6 ‚Äì SWM (Reinigung, Reklame‚Ä¶)
    "Reiniger":             (6, "SWM"),
    "Reinigung":            (6, "SWM"),
    "Reinigung Sa":         (6, "SWM"),
    "Reklame":              (6, "SWM"),
    "Sa Reklame":           (6, "SWM"),

    # 7 ‚Äì Unfall & Fenster
    "Lack":                 (7, "Unfall"),
    "Lackierer":            (7, "Unfall"),
    "Schaden":              (7, "Unfall"),
    "Scheibe":              (7, "Fenster"),
    "Speng":                (7, "Unfall"),
    "Spengler":             (7, "Unfall"),
    "Unfall":               (7, "Unfall"),

    # 8 ‚Äì SWM (Sonder-Fahrten, Schulung‚Ä¶)
    "Aufr√ºstung":           (8, "SWM"),
    "AW":                   (8, "SWM"),
    "BBH West":             (8, "SWM"),
    "Bhf Ost":              (8, "SWM"),
    "Bhf West":             (8, "SWM"),
    "Entwerter":            (8, "SWM"),
    "FA":                   (8, "SWM"),
    "Fahrschule":           (8, "SWM"),
    "Film":                 (8, "SWM"),
    "Foto":                 (8, "SWM"),
    "Funk":                 (8, "SWM"),
    "Piktogramme":          (8, "SWM"),
    "Pixida":               (8, "SWM"),
    "Schulung":             (8, "SWM"),
    "Sonderf.":             (8, "SWM"),
    "Sonderfahrt":          (8, "SWM"),
    "Stelzer":              (8, "SWM"),
    "Telematik":            (8, "SWM"),
    "Um-bau":               (8, "SWM"),
    "Vermietet":            (8, "SWM"),

    # 9 ‚Äì Fremdfirma
    "AOB":                  (9, "Fremdfirma"),
    "B u. M":               (9, "Fremdfirma"),
    "Beissbarth":           (9, "Fremdfirma"),
    "Edelwei√ü":             (9, "Fremdfirma"),
    "Fa Bode":              (9, "Fremdfirma"),
    "Fa Knorr":             (9, "Fremdfirma"),
    "Fa Lawo":              (9, "Fremdfirma"),
    "Fa Mayo":              (9, "Fremdfirma"),
    "Gr√§felfing":           (9, "Fremdfirma"),
    "Gruber":               (9, "Fremdfirma"),
    "HJS":                  (9, "Fremdfirma"),
    "H√ºttinger":            (9, "Fremdfirma"),
    "Kienzle":              (9, "Fremdfirma"),
    "Kronberger":           (9, "Fremdfirma"),
    "KWS":                  (9, "Fremdfirma"),
    "ML":                   (9, "Fremdfirma"),
    "√ñPNV":                 (9, "Fremdfirma"),
    "-":                  (None, ""),
    "Abmelden":           (None, "SWM"),
    "ITCS 2.0":           (None, "SWM"),
    "ITCS 2.0 (Hdy)":     (None, "SWM"),
    "Kein Plan":          (None, ""),
    "Nur Solo":           (None, ""),
    "o. Feiertag":        (None, ""),
    "St":                 (None, "SWM"),
    "Steht":              (None, "SWM"),
    "Wochenende":         (None, ""),
    "XXX":                (None, "SWM"),
    # ‚ÄöEinsatz i.O.‚Äò und √§hnliche
    "Einsatz i.O.":       (None, "SWM"),
    "geliefert":          (None, "SWM"),      # OOO in Deiner Liste
    "Nicht geliefert":    (None, "SWM"),      # XXX in Deiner Liste
    "Abnahme":            (None, "SWM"),      # ZZZ in Deiner Liste
    
    
}                       


# 1) Einmaliges Einlesen der Excel-Datei mit BusNr ‚Üí Hersteller
BASE_DIR=os.path.dirname(__file__)
EXCEL_PATH = os.path.join(BASE_DIR, "bus_hersteller_zuordnung.xlsx")

mapping_df = pd.read_excel(EXCEL_PATH, engine="openpyxl")
# Passe die √úberschriften an, falls deine Excel-Cols anders hei√üen:
mapping_df.columns = ["BusNr", "Hersteller"]
# Erstelle das Dict
BUS_TO_HERSTELLER: Dict[int, str] = mapping_df.set_index("BusNr")["Hersteller"].to_dict()

# ---------------------------------------------------------------------------
# 1) Seiteneinstellungen
# -----------------------------------------------------------------------------
def setup_page(
    title: str = "Ausfall‚ÄêAnalyse Busflotte",
    layout: str = "wide",
    sidebar_state: str = "expanded"
) -> None:
    st.set_page_config(page_title=title, layout=layout, initial_sidebar_state=sidebar_state)


def prepare_filtered_summary(
    summary_path: Path = Path(__file__).parent / "Zusammenfassung.xlsx",
    date_path:    Path = Path(__file__).parent / "Zulassung-Verkauf.xlsx",
    output_path:  Path = Path(__file__).parent / "Zusammenfassung_bearbeitet.xlsx",
    sheet_dates                  = 0
) -> None:
    """
    Schreibt in output_path zwei Sheets: 'Osten' und 'Moosach'.
    Jeder Sheet wird:
      1) aus summary_path eingelesen,
      2) in Long-Format gemolten und nach Zulassung/Verkauf
         gefiltert (au√üerhalb-Zeilen komplett entfernt),
      3) zur√ºck ins Wide-Format gepivottet (Datum x BusNr),
      4) unter demselben Sheet-Namen wieder ausgegeben.
    """

    # Alte Ausgabe l√∂schen
    if os.path.exists(output_path):
        os.remove(output_path)
    BUS_MAP_FILE = Path(__file__).parent / "bus_hersteller_zuordnung.xlsx"
    mapping_df = pd.read_excel(BUS_MAP_FILE, engine="openpyxl")
    # Zulassungs-/Verkaufs-Daten einlesen
    df_dates = pd.read_excel(date_path, sheet_name=sheet_dates, engine="openpyxl")
    df_dates.columns = df_dates.columns.str.strip()
    df_dates = df_dates.rename(columns={
        "KOM-Nr.":   "BusNr",
        "Einsatz":   "ZulassungDatum",
        "Verkauf":   "VerkaufDatum"
    })
    # Sicherstellen, dass die n√∂tigen Spalten da sind
    if not {"BusNr","ZulassungDatum"}.issubset(df_dates.columns):
        st.error("In der Zulassungs-Datei fehlen KOM-Nr. oder Zulassung.")
        return
    if "VerkaufDatum" not in df_dates.columns:
        df_dates["VerkaufDatum"] = pd.NaT

    # Datentypen korrigieren
    df_dates["BusNr"]          = df_dates["BusNr"].astype(str).str.strip()
    df_dates["ZulassungDatum"] = pd.to_datetime(df_dates["ZulassungDatum"], dayfirst=False, errors="coerce")
    df_dates["VerkaufDatum"]   = pd.to_datetime(df_dates["VerkaufDatum"],   dayfirst=False, errors="coerce")

    bereiche = ["Osten","Moosach"]
    result_wides = {}

    for bereich in bereiche:
        try:
            df_wide = pd.read_excel(summary_path, sheet_name=bereich, engine="openpyxl")
        except ValueError:
            st.warning(f"Sheet '{bereich}' nicht gefunden ‚Äì schreibe leeres Blatt.")
            result_wides[bereich] = pd.DataFrame({"Datum": []})
            continue

        # Spalten s√§ubern und Datum parsen
        df_wide.columns = df_wide.columns.str.strip()
        df_wide["Datum"] = pd.to_datetime(df_wide["Datum"], dayfirst=True, errors="coerce")

        # Long-Melt
        bus_cols = [c for c in df_wide.columns if c != "Datum"]
        df_long = df_wide.melt(
            id_vars=["Datum"],
            value_vars=bus_cols,
            var_name="BusNr",
            value_name="Ausfallgrund"
        )
        df_long["BusNr"] = df_long["BusNr"].astype(str).str.strip()

        # Merge mit s-/Verkaufsdaten
        df_long = df_long.merge(
            df_dates[["BusNr","ZulassungDatum","VerkaufDatum"]],
            on="BusNr", how="left"
        )

        # Maske: Datum im g√ºltigen Bereich?
        mask = (
            (df_long["Datum"] >= df_long["ZulassungDatum"]) &
            (
                df_long["VerkaufDatum"].isna() |
                (df_long["Datum"] <= df_long["VerkaufDatum"])
            )
        ).fillna(False)

        # Nur g√ºltige Zeilen behalten (alles au√üerhalb wird entfernt)
        df_long = df_long[mask].copy()

        # Duplikate entfernen (falls mehrfach gemeldet)
        df_long = (
            df_long
            .sort_values(["Datum","BusNr"])
            .drop_duplicates(subset=["Datum","BusNr"], keep="last")
        )

        # Zur√ºck ins Wide-Format pivotieren
        df_new_wide = df_long.pivot(
            index="Datum",
            columns="BusNr",
            values="Ausfallgrund"
        ).reset_index()
        df_new_wide.columns.name = None

        # Bus-Spalten sauber sortieren
        bus_spalten = [c for c in df_new_wide.columns if c != "Datum"]
        bus_spalten = sorted(
            bus_spalten,
            key=lambda x: (0, int(x)) if x.isdigit() else (1, x)
        )
        df_new_wide = df_new_wide[["Datum"] + bus_spalten]

        result_wides[bereich] = df_new_wide

    # Ergebnis in zwei sichtbare Sheets schreiben
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for bereich, df_out in result_wides.items():
            df_out.to_excel(writer, sheet_name=bereich, index=False)

    size = os.path.getsize(output_path)
    st.success(f"‚úÖ Gefilterte Datei geschrieben: {output_path} ({size} Bytes)")
# -----------------------------------------------------------------------------
@st.cache_data
def get_data(path: str) -> pd.DataFrame:
    df = load_data(path)
    # Stelle sicher, dass 'Datum' als datetime vorliegt
    df["Datum"] = pd.to_datetime(df["Datum"])
    return df

def _load_registration_dates(
    date_path = Path(__file__).parent/"Zulassung-Verkauf.xlsx",
    sheet_dates: int = 0
) -> pd.DataFrame:
    """
    Liefert DataFrame mit Spalten: BusNr, ZulassungDatum, VerkaufDatum
    """
    df_dates = pd.read_excel(date_path, sheet_name=sheet_dates, engine="openpyxl")
    df_dates.columns = df_dates.columns.str.strip()
    df_dates = df_dates.rename(columns={
        "KOM-Nr.": "BusNr",
        "Einsatz": "ZulassungDatum",
        "Verkauf": "VerkaufDatum"
    })
    df_dates["BusNr"] = df_dates["BusNr"].astype(str).str.strip()
    df_dates["ZulassungDatum"] = pd.to_datetime(df_dates["ZulassungDatum"], errors="coerce")
    # falls Verkauf fehlt, als NaT setzen
    if "VerkaufDatum" in df_dates.columns:
        df_dates["VerkaufDatum"] = pd.to_datetime(df_dates["VerkaufDatum"], errors="coerce")
    else:
        df_dates["VerkaufDatum"] = pd.NaT
    return df_dates
def plot_series_status_heatmap(
    df_filt: pd.DataFrame,
    df_dates: pd.DataFrame,
    dpi: int = 200,
    figsize: tuple = (16.53, 11.69),  # A3 Landscape in inches
    filename: str = None
) -> None:
    # --- 1) Status-Spalte sicherstellen ---
    df = df_filt.copy()
    if "Status" not in df.columns:
        if "Ausfall-Typ" in df.columns:
            df["Status"] = df["Ausfall-Typ"].map(
                lambda x: "Ausgefallen" if x != "Fahren" else "Fahren"
            )
        else:
            raise KeyError("df_filt muss 'Status' oder 'Ausfall-Typ' enthalten.")
    df["Datum"] = pd.to_datetime(df["Datum"])

    # --- 2) Zulassungs-/Verkaufsdaten einlesen und filtern ---
    dates = df_dates.copy()
    dates["BusNr"] = dates["BusNr"].astype(str).str.strip()
    dates["ZulassungDatum"] = pd.to_datetime(dates["ZulassungDatum"], errors="coerce")
    dates["VerkaufDatum"]   = pd.to_datetime(dates.get("VerkaufDatum"), errors="coerce")

    start, end = df["Datum"].min(), df["Datum"].max()
    all_dates  = pd.date_range(start, end, freq="D")
    buses = (
        df[["BusNr","Serie"]]
        .drop_duplicates()
        .sort_values(["Serie","BusNr"])
        .reset_index(drop=True)
    )

    # Vollmatrix Bus√óDatum
    full = (
        pd.MultiIndex
          .from_product([buses["BusNr"], all_dates], names=["BusNr","Datum"])
          .to_frame(index=False)
    )
    full = full.merge(buses, on="BusNr", how="left")
    full = full.merge(dates, on="BusNr", how="left")

    # Maske: nur Service-Tage (zwischen Zulassung und Verkauf)
    mask_service = (
        (full["Datum"] >= full["ZulassungDatum"]) &
        ((full["VerkaufDatum"].isna()) | (full["Datum"] <= full["VerkaufDatum"]))
    ).fillna(False)
    full = full.loc[mask_service, ["BusNr","Datum","Serie"]]

    # Merge mit dem echten Status (Ausfall / Fahren)
    full = full.merge(
        df[["BusNr","Datum","Status"]],
        on=["BusNr","Datum"],
        how="left"
    )
    full["Status"] = full["Status"].fillna("Fahren")

    # --- 3) Pivot mit pivot_table und Aggfunc ---
    pivot = full.pivot_table(
        index="BusNr",
        columns="Datum",
        values="Status",
        aggfunc="first"      # falls trotzdem Duplikate da sind, nimm den ersten
    )

    # Stelle sicher, dass die Zeilen in der Reihenfolge buses["BusNr"] stehen
    pivot = pivot.reindex(index=buses["BusNr"])

    # Numerische Matrix: 0=Fahren, 1=Ausgefallen, NaN=kein Service
    Z = pivot.replace({"Fahren": 0, "Ausgefallen": 1}).astype(float).values

    # --- 4) Bus‚ÜíSerie Mapping in Pivot-Reihenfolge ---
    bus_to_serie = buses.set_index("BusNr")["Serie"].to_dict()
    series_per_row = [bus_to_serie[bus] for bus in pivot.index]

    # Grenzen zwischen Serienbl√∂cken finden
    borders = [
        i - 0.5
        for i in range(1, len(series_per_row))
        if series_per_row[i] != series_per_row[i-1]
    ]

    # Mittelpunkte f√ºr Y-Ticks je Serie
    tick_pos   = []
    tick_label = []
    start_i = 0
    for i in range(1, len(series_per_row)+1):
        end_i = i-1
        if (i == len(series_per_row)) or (series_per_row[i] != series_per_row[i-1]):
            center = (start_i + end_i) / 2
            tick_pos.append(center)
            tick_label.append(series_per_row[end_i])
            start_i = i

    # --- 5) Plot auf A3-Gr√∂√üe ---
    fig, ax = plt.subplots(figsize=figsize, dpi=dpi)
    cmap = ListedColormap(["#00AA00", "#DD3333"])
    cmap.set_bad(color="#EEEEEE")  # Service-loses ‚Üí hellgrau
    ax.imshow(Z, aspect="auto", cmap=cmap, origin="lower")

    # Y-Linien zwischen Serien
    for y in borders:
        ax.axhline(y=y, color="white", linewidth=2)

    ax.set_yticks(tick_pos)
    ax.set_yticklabels(tick_label, fontsize=8)
    ax.set_ylabel("Serie")

    # --- 6) Monats- & Jahreslinien auf X ---
    dates_idx    = pivot.columns
    month_bounds = [i for i, d in enumerate(dates_idx) if d.day == 1]
    year_bounds  = [i for i, d in enumerate(dates_idx) if (d.day == 1 and d.month == 1)]

    for mb in month_bounds:
        ax.axvline(x=mb-0.5, color="white", linewidth=0.8)
    for yb in year_bounds:
        ax.axvline(x=yb-0.5, color="black", linewidth=1.5)

    # X-Ticks nur an Monatsanf√§ngen
    ax.set_xticks(month_bounds)
    labels = []
    for idx in month_bounds:
        d = dates_idx[idx]
        lbl = d.strftime("%b")
        if idx in year_bounds:
            lbl += f"\n{d.year}"
        labels.append(lbl)
    ax.set_xticklabels(labels, rotation=90, fontsize=7)
    ax.set_xlabel("Monat / Jahr")

    # --- 7) Legende ---
    from matplotlib.patches import Patch
    legend_handles = [
        Patch(facecolor="#00AA00", label="Fahren"),
        Patch(facecolor="#DD3333", label="Ausgefallen"),
        Patch(facecolor="#EEEEEE", label="kein Service")
    ]
    ax.legend(
        handles=legend_handles,
        loc="upper left",
        bbox_to_anchor=(1.01, 1),
        fontsize=8
    )

    ax.set_title(
        f"Status-Heatmap (gr√ºn=Fahren, rot=Ausgefallen)\n"
        f"Zeitraum: {dates_idx[0].date()} ‚Äì {dates_idx[-1].date()}",
        fontsize=10,
        pad=12
    )
    plt.tight_layout()

    # --- 8) Speichern und in Streamlit ausgeben ---
    if filename:
        fig.savefig(filename, dpi=dpi)
    buf = BytesIO()
    fig.savefig(buf, format="png", dpi=dpi)
    buf.seek(0)
    st.image(buf, use_column_width=True)
    plt.close(fig)
    
    
    
def plot_status_heatmap(
    df_filt: pd.DataFrame,
    date_path = Path(__file__).parent/"Zulassung-Verkauf.xlsx",
    sheet_dates: int = 0
) -> None:
    """
    Erzeugt eine Heatmap: pro Bus und Datum, gr√ºn f√ºr 'Fahren', rot f√ºr 'Ausgefallen'.
    Vorab werden alle Tage au√üerhalb Zulassung/Verkauf komplett ignoriert.
    """
    import streamlit as st
    import pandas as pd
    import plotly.express as px

    if df_filt.empty:
        st.info("Keine Daten f√ºr Heatmap vorhanden.")
        return

    # 1) Zulassungs-/Verkaufsdaten laden
    df_dates = pd.read_excel(date_path, sheet_name=sheet_dates, engine="openpyxl")
    df_dates = df_dates.rename(columns={
        "KOM-Nr.": "BusNr",
        "Einsatz": "ZulassungDatum",
        "Verkauf": "VerkaufDatum"
    })
    df_dates["BusNr"] = df_dates["BusNr"].astype(str).str.strip()
    df_dates["ZulassungDatum"] = pd.to_datetime(df_dates["ZulassungDatum"], errors="coerce")
    if "VerkaufDatum" in df_dates.columns:
        df_dates["VerkaufDatum"] = pd.to_datetime(df_dates["VerkaufDatum"], errors="coerce")
    else:
        df_dates["VerkaufDatum"] = pd.NaT

    # 2) Globale Zeitspanne ermitteln
    start = df_filt["Datum"].min()
    end   = df_filt["Datum"].max()

    # 3) Vollmatrix aller Bus√óTage
    all_dates = pd.date_range(start, end, freq="D")
    all_buses = df_filt["BusNr"].unique()
    full = (
        pd.MultiIndex
          .from_product([all_buses, all_dates], names=["BusNr","Datum"])
          .to_frame(index=False)
    )

    # 4) Zulassungs-/Verkaufs-Filter anwenden
    full = full.merge(
        df_dates[["BusNr","ZulassungDatum","VerkaufDatum"]],
        on="BusNr", how="left"
    )
    mask = (
        (full["Datum"] >= full["ZulassungDatum"]) &
        (
            full["VerkaufDatum"].isna() |
            (full["Datum"] <= full["VerkaufDatum"])
        )
    ).fillna(False)
    full = full.loc[mask, ["BusNr","Datum"]]

    # 5) Ausgefallene Tage kennzeichnen
    failed = (
        df_filt[df_filt["Status"] == "Ausgefallen"]
        [["BusNr","Datum"]]
        .drop_duplicates()
        .assign(Status="Ausgefallen")
    )

    # 6) Merge und Default-Fahren
    status_df = (
        full
        .merge(failed, on=["BusNr","Datum"], how="left")
        .assign(Status=lambda d: d["Status"].fillna("Fahren"))
    )

    # 7) Drop-Duplicates und Pivot mit pivot_table
    status_df = status_df.drop_duplicates(subset=["BusNr","Datum"])
    pivot = status_df.pivot_table(
        index="BusNr",
        columns="Datum",
        values="Status",
        aggfunc="first"
    )

    # 8) Numerische Matrix f√ºr Plotly
    z = pivot.replace({"Fahren": 0, "Ausgefallen": 1}).values

    # 9) Plot
    fig = px.imshow(
        z,
        x=pivot.columns,
        y=pivot.index,
        color_continuous_scale=[(0, "green"), (1, "red")],
        aspect="auto",
        origin="lower"
    )
    fig.update_coloraxes(showscale=False)
    fig.update_layout(
        xaxis=dict(tickformat="%d.%m", tickangle=45),
        yaxis_title="BusNr",
        xaxis_title="Datum",
        margin=dict(l=50, r=20, t=30, b=80)
    )

    st.subheader("üóìÔ∏è Status-Heatmap (gr√ºn = gefahren, rot = ausgefallen)")
    st.plotly_chart(fig, use_container_width=True)

# -----------------------------------------------------------------------------
# 3) Sidebar‚ÄêFilter ‚Äì Pflicht‚ÄêMultiselects leer, st.stop() bei keiner Auswahl
# -----------------------------------------------------------------------------
def sidebar_filters(df: pd.DataFrame) -> Dict[str, Any]:
    st.sidebar.markdown("## üîé Filter")

    # ‚Äî 1) Zeit-Filter (Datum vs. Quartal) ‚Äî
    zeitwahl = st.sidebar.radio("Nach welchem Zeitraum filtern?", ["Datum", "Quartal"], index=0)
    if zeitwahl == "Datum":
        min_d, max_d = df["Datum"].min(), df["Datum"].max()
        datum_start, datum_ende = st.sidebar.date_input(
            "Datum von‚Äìbis",
            value=[min_d, max_d],
            min_value=min_d,
            max_value=max_d
        )
        quartal = None
    else:
        datum_start, datum_ende = None, None
        quartal = st.sidebar.multiselect(
            "Quartal",
            options=sorted(df["Jahr-Quartal"].unique()),
            default=[]
        )
        if not quartal:
            st.sidebar.info("Bitte w√§hle mindestens ein Quartal aus.")
            st.stop()

    # ‚Äî 2) Hersteller-Spalte erg√§nzen, falls noch nicht da ‚Äî
    if "Hersteller" not in df.columns:
        df = df.copy()
        df["Hersteller"] = df["BusNr"].map(BUS_TO_HERSTELLER).fillna("Unbekannt")

    # ‚Äî 3) Filter nach Busnummer / Busserie / Hersteller ‚Äî
    buswahl = st.sidebar.radio("Filter nach", ["Busnummer", "Busserie", "Hersteller"], index=0)
    if buswahl == "Busnummer":
        all_buses = st.sidebar.checkbox("Alle Busnummern ausw√§hlen", key="all_buses")
        bus_options = sorted(df["BusNr"].unique())
        default_buses = bus_options if all_buses else []
        busnr = st.sidebar.multiselect("Busnummer(n)", options=bus_options, default=default_buses)
        serie = None
        hersteller = None
        if not busnr:
            st.sidebar.info("Bitte w√§hle mindestens eine Busnummer aus.")
            st.stop()
    elif buswahl == "Busserie":
        serie = st.sidebar.multiselect("Busserie(n)", options=sorted(df["Serie"].unique()), default=[])
        busnr = None
        hersteller = None
        if not serie:
            st.sidebar.info("Bitte w√§hle mindestens eine Serie aus.")
            st.stop()
    else:
        all_hers = st.sidebar.checkbox("Alle Hersteller ausw√§hlen", key="all_hers")
        hers_opts = sorted(df["Hersteller"].unique())
        default_hers = hers_opts if all_hers else []
        hersteller = st.sidebar.multiselect("Hersteller", options=hers_opts, default=default_hers)
        busnr = None
        serie = None
        if not hersteller:
            st.sidebar.info("Bitte w√§hle mindestens einen Hersteller aus.")
            st.stop()

    # ‚Äî 4) Ausfall-Typ ‚Äî
    typ = st.sidebar.multiselect(
        "Ausfall-Typ",
        options=["Standtage", "Einr√ºcker", "Sonstiges", "Fahren"],
        default=["Standtage", "Einr√ºcker", "Sonstiges", "Fahren"]
    )

    # ‚Äî 5) Auswahl-Context zum Einschr√§nken der Ausfallgrund‚ÄêOptionen vorbereiten ‚Äî
    #    Wir erstellen eine Maske genau wie in filter_and_add_km, ohne km‚ÄêSpalten:
    df_opt = df.copy()
    mask = pd.Series(True, index=df_opt.index)

    # 5A) Zeit-Filter
    if datum_start is not None and datum_ende is not None:
        start = pd.to_datetime(datum_start)
        ende  = pd.to_datetime(datum_ende)
        mask &= (df_opt["Datum"] >= start) & (df_opt["Datum"] <= ende)
    else:
        mask &= df_opt["Jahr-Quartal"].isin(quartal)

    # 5B) Bus-Filter
    if busnr:
        mask &= df_opt["BusNr"].isin(busnr)
    elif serie:
        mask &= df_opt["Serie"].isin(serie)
    elif hersteller:
        mask &= df_opt["Hersteller"].isin(hersteller)

    # 5C) Ausfall-Typ
    if "Ausfall-Typ" in df_opt.columns and typ:
        mask &= df_opt["Ausfall-Typ"].isin(typ)

    df_opt = df_opt[mask]

    # ‚Äî 6) Ausfallgrund(e) mit H√§ufigkeit in Klammern ‚Äî
    if "Ausfallgrund" not in df_opt.columns or df_opt["Ausfallgrund"].dropna().empty:
        st.sidebar.info("Keine Ausfallgr√ºnde vorhanden f√ºr die gew√§hlten Filter.")
        selected_gr = []
        label_to_gr = {}
    else:
        # H√§ufigkeiten berechnen
        gr_counts = df_opt["Ausfallgrund"].value_counts(dropna=True)
        # Labels "Grund (N)"
        options = [f"{gr} ({gr_counts[gr]})" for gr in gr_counts.index]
        # Mapping zur√ºck auf den Original‚ÄìGrund
        label_to_gr = {opt: gr for opt, gr in zip(options, gr_counts.index)}
        # Multiselect
        selected_labels = st.sidebar.multiselect(
            "Ausfallgrund(e) ausw√§hlen",
            options=options,
            default=options
        )
        # zur√ºck√ºbersetzen
        selected_gr = [label_to_gr[label] for label in selected_labels]

    # ‚Äî 7) Restliche Sidebar‚ÄêEinstellungen ‚Äî
    top_n         = st.sidebar.slider("Top N Ausfallgr√ºnde im Pie", 3, 15, 7)
    zeit_gruppe   = st.sidebar.radio("Zeit gruppieren nach", ["T√§glich", "W√∂chentlich", "Monatlich"])
    ts_typ        = st.sidebar.selectbox("Typ Zeitreihe", ["Linie", "Fl√§che", "Balken"])
    diskret       = st.sidebar.selectbox("Diskretes Farbschema", list(DISCRETE_SCHEMAS.keys()), index=0)
    kontinuierlich = st.sidebar.selectbox("Kontinuierliches Farbschema", list(CONTINUOUS_SCHEMAS.keys()), index=0)

    st.sidebar.markdown("## ‚öôÔ∏è Standard-Kilometer pro Typ")
    einr_km   = st.sidebar.number_input("Default km Einr√ºcker",  min_value=0, value=50,  step=10)
    stand_km  = st.sidebar.number_input("Default km Standtage",  min_value=0, value=0,   step=10)
    son_km    = st.sidebar.number_input("Default km Sonstiges",  min_value=0, value=250, step=10)
    fahr_km   = st.sidebar.number_input("Default km Fahren",     min_value=0, value=250, step=10)

    return dict(
        datum_start    = datum_start,
        datum_ende     = datum_ende,
        quartal        = quartal,
        busnr          = busnr,
        serie          = serie,
        hersteller     = hersteller,
        typ            = typ,
        ausfallgrund   = selected_gr,
        top_n          = top_n,
        zeit_gruppe    = zeit_gruppe,
        ts_typ         = ts_typ,
        diskret        = diskret,
        kontinuierlich = kontinuierlich,
        km_defaults    = {"Einr√ºcker": einr_km, "Standtage": stand_km, "Sonstiges": son_km},
        km_fahren      = fahr_km
    )


# -----------------------------------------------------------------------------
# 4) Daten filtern und km‚ÄêSpalte erg√§nzen + DataEditor
# -----------------------------------------------------------------------------
def filter_and_add_km(
    df: pd.DataFrame,
    filt: Dict[str, Any]
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Filtert das DataFrame nach Datum/Quartal, BusNr/Serie/Hersteller und Ausfall-Typ.
    F√ºgt die km-Spalten hinzu (km_default, km_fahren, km).
    Entfernt nicht mehr Tage vor dem ersten Ausfall.
    """
    df2 = df.copy()
    mask = pd.Series(True, index=df2.index)

    # 1) Status-Spalte sicherstellen
    if "Status" not in df2.columns:
        if "Ausfall-Typ" in df2.columns:
            df2["Status"] = df2["Ausfall-Typ"].apply(
                lambda x: "Ausgefallen" if x != "Fahren" else "Fahren"
            )
        else:
            raise KeyError("Weder 'Status' noch 'Ausfall-Typ' in den Daten gefunden.")

    # 2) Filter Datum oder Quartal
    if filt["datum_start"] and filt["datum_ende"]:
        start = pd.to_datetime(filt["datum_start"])
        ende  = pd.to_datetime(filt["datum_ende"])
        mask &= (df2["Datum"] >= start) & (df2["Datum"] <= ende)
    else:
        mask &= df2["Jahr-Quartal"].isin(filt["quartal"])

    # 3) Filter BusNr / Serie / Hersteller
    if filt["busnr"]:
        mask &= df2["BusNr"].isin(filt["busnr"])
    elif filt["serie"]:
        mask &= df2["Serie"].isin(filt["serie"])
    elif filt["hersteller"]:
        mask &= df2["Hersteller"].isin(filt["hersteller"])

    # 4) Filter Ausfall-Typ
    if "Ausfall-Typ" in df2.columns and filt["typ"]:
        mask &= df2["Ausfall-Typ"].isin(filt["typ"])

    # 5) Anwenden aller Filter
    df_filt = df2[mask].copy()

    # 6) km-Logik
    if "Ausfallgrund" in df_filt.columns:
        df_filt = df_filt[df_filt["Ausfallgrund"] != "Keine Ausf√§lle"]
    df_filt["km_default"] = df_filt["Ausfall-Typ"].map(filt["km_defaults"])
    df_filt["km_fahren"]  = filt["km_fahren"]
    df_filt["km"]         = df_filt["km_default"]

    # F√ºr die KM-Auswertung behalten wir eine Kopie
    df_km = df_filt.copy()
    return df_filt, df_km



# -----------------------------------------------------------------------------
# 5) Analyse‚ÄêSeite
# -----------------------------------------------------------------------------
def page_analyse(
    df_filt: pd.DataFrame,
    df_km:   pd.DataFrame,         # neu hinzugef√ºgt
    filt:    Dict[str, Any]
) -> None:
    st.title("üöç Ausfall‚ÄêAnalyse")

    # KPI‚ÄêLeiste
    tage = max(df_filt["Datum"].nunique(), 1)
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Zeitraum", f"{filt['datum_start']} bis {filt['datum_ende']}")
    c2.metric("Quartale", ", ".join(filt["quartal"]))
    c3.metric("Ausf√§lle gesamt", len(df_filt))
    c4.metric("√ò Ausf√§lle/Tag", f"{len(df_filt)/tage:.2f}")
    st.markdown("---")

    # Aggregationsebene
    group_by = st.selectbox(
        "W√§hle die Aggregationsebene:",
        options=["Ausfallgrund", "BusNr", "Serie", "Jahr-Quartal"]
    )
    counts = (
        df_filt[group_by]
        .value_counts()
        .reset_index(name="Anzahl")
        .rename(columns={"index": group_by})
    )
    fig = px.bar(
        counts,
        x=group_by,
        y="Anzahl",
        title=f"Ausf√§lle nach {group_by}",
        color=group_by if group_by in ["Ausfallgrund", "Serie"] else None,
        color_discrete_sequence=DISCRETE_SCHEMAS[filt["diskret"]] if group_by in ["Ausfallgrund", "Serie"] else None
    )
    st.plotly_chart(fig, use_container_width=True)
    st.markdown("---")

    # Detail‚ÄêCharts Serie √ó Ausfallgrund
    st.subheader("Serie √ó Ausfallgrund")
    ser = st.selectbox("Serie ausw√§hlen", sorted(df_filt["Serie"].unique()))
    grd = st.selectbox("Ausfallgrund ausw√§hlen", sorted(df_filt["Ausfallgrund"].unique()))
    df_detail = df_filt[(df_filt["Serie"] == ser) & (df_filt["Ausfallgrund"] == grd)]
    st.write(f"Anzahl Ausf√§lle in Serie **{ser}** mit Grund **{grd}**: **{len(df_detail)}**")
    if not df_detail.empty:
        df_time = df_detail.set_index("Datum").resample("W").size().reset_index(name="Anzahl")
        fig_detail = px.line(
            df_time, x="Datum", y="Anzahl",
            title=f"W√∂chentliche Ausf√§lle ‚Äì Serie {ser}, Grund {grd}",
            markers=True,
            color_discrete_sequence=[DISCRETE_SCHEMAS[filt["diskret"]][0]]
        )
        st.plotly_chart(fig_detail, use_container_width=True)
    else:
        st.info("Keine Daten f√ºr diese Kombination vorhanden.")


# -----------------------------------------------------------------------------
# 6) Statistik‚ÄêSeite
# -----------------------------------------------------------------------------
def page_statistik(
    df_filt: pd.DataFrame,
    df_km: pd.DataFrame,
    km_fahren: int,
    kontinuierlich: str
) -> None:
    st.title("üìä Grundstatistik & KM‚ÄêBetrachtung")
    if df_filt.empty:
        st.warning("Keine Daten f√ºr die ausgew√§hlten Filter.")
        st.stop()

    # H√§ufigkeiten
    for name, col in [("Ausfallgr√ºnde","Ausfallgrund"),("Busse","BusNr"),("Serien","Serie")]:
        st.subheader(name)
        tab = df_filt[col].value_counts().reset_index(name="Anzahl").rename(columns={"index":col})
        st.dataframe(tab, use_container_width=True)

    # Pivot‚ÄêTabelle
    st.markdown("### Pivot‚ÄêTabelle (Serie √ó Ausfallgrund)")
    pivot = df_filt.pivot_table(index="Serie", columns="Ausfallgrund", aggfunc="size", fill_value=0)
    st.dataframe(pivot, use_container_width=True)

    # Quartal‚ÄêDiagramm
    quart = (df_filt["Jahr-Quartal"]
             .value_counts()
             .reset_index(name="Anzahl")
             .rename(columns={"index":"Jahr-Quartal"})
             .sort_values("Jahr-Quartal"))
    st.markdown("### Ausf√§lle pro Quartal")
    fig_q = px.bar(
        quart, x="Jahr-Quartal", y="Anzahl",
        color="Anzahl",
        color_continuous_scale=CONTINUOUS_SCHEMAS[kontinuierlich]
    )
    st.plotly_chart(fig_q, use_container_width=True)

    # KM‚ÄêAuswertung pro Bus
    st.markdown("### üõ£Ô∏è KM‚ÄêAuswertung pro Bus")
    bus_km = (
        df_km.groupby("BusNr")
             .agg(Tage=("Datum","nunique"), km_ist=("km","sum"))
             .reset_index()
    )
    bus_km["km_soll"] = bus_km["Tage"] * km_fahren
    bus_km["Verf_%"]  = (bus_km["km_ist"] / bus_km["km_soll"] * 100).round(1)
    st.dataframe(bus_km, use_container_width=True)

    # Rohdaten‚ÄêExport
    excel_bytes = to_excel_raw(df_filt)
    st.download_button(
        "üì• Rohdaten als Excel herunterladen",
        data=excel_bytes,
        file_name="rohdaten_export.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# -----------------------------------------------------------------------------
# 7) KM‚ÄêBetrachtung‚ÄêSeite
# -----------------------------------------------------------------------------
def page_km_betrachtung(
    df_filt: pd.DataFrame,
    df_km: pd.DataFrame,
    km_fahren: int
) -> None:
    st.title("üõ£Ô∏è KM-Betrachtung")
    if df_km.empty:
        st.warning("Keine KM-Daten vorhanden.")
        st.stop()

    # Vollmatrix
    all_dates = pd.date_range(df_km["Datum"].min(), df_km["Datum"].max(), freq="D")
    all_buses = df_km["BusNr"].unique()
    full = (
        pd.MultiIndex.from_product([all_dates, all_buses], names=["Datum","BusNr"])
          .to_frame(index=False)
          .merge(df_km, on=["Datum","BusNr"], how="left")
    )
    # finden, wie die Typ-Spalte hei√üt
    type_col = "Ausfall-Typ" if "Ausfall-Typ" in full.columns else "Typ"
    full[type_col] = full[type_col].fillna("Fahren")
    full["km"]  = full["km"].fillna(km_fahren)
    serie_map = df_filt[["BusNr","Serie"]].drop_duplicates()
    full = full.merge(serie_map, on="BusNr", how="left")

    # Histogramm
    st.subheader("Verteilung der gefahrenen km je Typ")
    fig_dist = px.histogram(
    full,
    x="km",
    color=type_col,
    nbins=30,
    title="Histogramm km-Werte pro Typ",
    color_discrete_sequence=px.colors.qualitative.Plotly
)
    st.plotly_chart(fig_dist, use_container_width=True)

    # Zusammenfassung pro Bus
    st.subheader("km-Zusammenfassung pro Bus")
    bus_km2 = (
        full.groupby("BusNr")
            .agg(Tage=("Datum","nunique"), km_ist=("km","sum"))
            .reset_index()
    )
    bus_km2["km_soll"] = bus_km2["Tage"] * km_fahren
    bus_km2["Verf_%"]  = (bus_km2["km_ist"] / bus_km2["km_soll"] * 100).round(1)
    st.dataframe(bus_km2, use_container_width=True)

    # Charts
    col1, col2 = st.columns(2)
    with col1:
        fig_ist = px.bar(
            bus_km2, x="BusNr", y="km_ist",
            title="Ist-Kilometer pro Bus",
            color="BusNr", color_continuous_scale="Viridis"
        )
        st.plotly_chart(fig_ist, use_container_width=True)
    with col2:
        fig_verf = px.bar(
            bus_km2, x="BusNr", y="Verf_%",
            title="Verf√ºgbarkeit in %",
            color="Verf_%", color_continuous_scale="Magma"
        )
        st.plotly_chart(fig_verf, use_container_width=True)

    # Serie‚ÄêZusammenfassung
    st.subheader("km-Zusammenfassung pro Bus-Serie")
    serie_km = (
        bus_km2.merge(serie_map, on="BusNr")
               .groupby("Serie")
               .agg(km_ist=("km_ist","sum"), km_soll=("km_soll","sum"))
               .reset_index()
    )
    serie_km["Verf_%"] = (serie_km["km_ist"] / serie_km["km_soll"] * 100).round(1)
    fig_serie = px.bar(
        serie_km, x="Serie", y="Verf_%",
        title="Verf√ºgbarkeit pro Serie (%)",
        color="Verf_%", color_continuous_scale="Plasma"
    )
    st.plotly_chart(fig_serie, use_container_width=True)
    

def page_kategorien(
    df_filt: pd.DataFrame,
    diskret: str
) -> None:
    st.title("üè∑Ô∏è Auswertung nach Fehlerkategorien")

    if df_filt.empty:
        st.warning("Es sind keine Daten f√ºr die ausgew√§hlten Filter verf√ºgbar.")
        st.stop()

    # Spalte f√ºr Ausfall-Typ ermitteln
    if "Ausfall-Typ" in df_filt.columns:
        typ_spalte = "Ausfall-Typ"
    elif "Typ" in df_filt.columns:
        typ_spalte = "Typ"
    else:
        st.error("Die Spalte f√ºr den Ausfall-Typ ('Ausfall-Typ' oder 'Typ') fehlt.")
        st.stop()

    # Nur echte Ausf√§lle (nicht Fahren) ‚Üí df_ausfall
    if "Ausfall" in df_filt.columns:
        df_ausfall = df_filt[df_filt["Ausfall"]].copy()
    else:
        df_ausfall = df_filt[df_filt[typ_spalte] != "Fahren"].copy()

    if df_ausfall.empty:
        st.info("Es gibt keine Ausf√§lle (Ausfall-Typ != 'Fahren') in den aktuellen Daten.")
        st.stop()

    # 1) Automatische Erkennung der Fehlerkategorie
    fehler_mapping = FEHLERKATEGORIEN  # oder auto_cat_name, je nachdem was Du brauchst

    df_ausfall["Fehlerkategorie"] = (
    df_ausfall["Ausfallgrund"]
      .replace(fehler_mapping)        # ersetzt jeden Schl√ºssel durch den Wert
      .fillna("Sonstiges")            # alles, was nicht matched, bekommt "Sonstiges"
)

    # 2) √úbersichtstabelle
    st.subheader("√úbersicht der Fehlerkategorien")
    fehler_counts = (
        df_ausfall["Fehlerkategorie"]
        .value_counts()
        .reset_index(name="Anzahl")
        .rename(columns={"index": "Fehlerkategorie"})
    )
    st.dataframe(fehler_counts, use_container_width=True)

    # 3) Balkendiagramm
    st.subheader("H√§ufigkeiten der Fehlerkategorien")
    fig_bar = px.bar(
        fehler_counts,
        x="Fehlerkategorie",
        y="Anzahl",
        color="Fehlerkategorie",
        color_discrete_sequence=DISCRETE_SCHEMAS[diskret],
        title="Anzahl der Ausf√§lle nach Fehlerkategorie"
    )
    fig_bar.update_layout(xaxis_tickangle=45)
    st.plotly_chart(fig_bar, use_container_width=True)

    # 4) Zeitliche Entwicklung
    st.subheader("Zeitliche Entwicklung der Fehlerkategorien")
    gruppierung = st.selectbox("Zeitliche Gruppierung", ["W√∂chentlich", "Monatlich"], index=0)
    freq = "W" if gruppierung == "W√∂chentlich" else "M"

    # Gruppieren nach Periode + Kategorie
    zeitliche = (
        df_ausfall
        .groupby([
            pd.Grouper(key="Datum", freq=freq),
            "Fehlerkategorie"
        ])
        .size()
        .reset_index(name="Anzahl")
    )

    # Damit auch Perioden ohne Ausfall einer Kategorie mit 0 auftauchen,
    # erzeugen wir ein vollst√§ndiges Raster aus allen Datums¬≠punkten √ó Kategorien:
    all_dates = pd.date_range(
        start=zeitliche["Datum"].min(),
        end=zeitliche["Datum"].max(),
        freq=freq
    )
    all_cats = zeitliche["Fehlerkategorie"].unique()
    full_index = pd.MultiIndex.from_product(
        [all_dates, all_cats],
        names=["Datum", "Fehlerkategorie"]
    )
    zeitliche = (
        zeitliche
        .set_index(["Datum", "Fehlerkategorie"])
        .reindex(full_index, fill_value=0)
        .reset_index()
    )

    # Plot
    fig_area = px.area(
        zeitliche,
        x="Datum",
        y="Anzahl",
        color="Fehlerkategorie",
        color_discrete_sequence=DISCRETE_SCHEMAS[diskret],
        title=f"Entwicklung der Fehlerkategorien ({gruppierung})"
    )
    st.plotly_chart(fig_area, use_container_width=True)

    # 5) Top-N im Kuchen
    st.subheader("Top Fehlerkategorien")
    top_n = st.slider("Anzahl der Top-Kategorien", 3, 20, 7)
    top_kats = fehler_counts.head(top_n)
    fig_pie = px.pie(
        top_kats,
        names="Fehlerkategorie",
        values="Anzahl",
        color_discrete_sequence=DISCRETE_SCHEMAS[diskret],
        title=f"Kuchen-Diagramm der Top {top_n} Fehlerkategorien"
    )
    st.plotly_chart(fig_pie, use_container_width=True)
    
    
    
    

def page_uebersicht(df_filt: pd.DataFrame, filt: Dict[str, Any]) -> None:
    """
    Kombinierte √úbersichtsseite ohne KM-Betrachtung:
     - KPIs
     - Balkendiagramme nach Ausfallgrund und Bus
     - Top-N Ausfallgr√ºnde
     - Heatmap der Ausfall-/Fahrstatus
     - Fehlerkategorien-Auswertung
    """
    st.title("üìã Gesamtauswertung Ausf√§lle")

    # --- 1) KPIs ---
    tage   = max(df_filt["Datum"].nunique(), 1)
    gesamt = len(df_filt)
    avg_tag = gesamt / tage
    c1, c2, c3 = st.columns(3)
    c1.metric("Zeitraum", f"{filt['datum_start']} bis {filt['datum_ende']}")
    c2.metric("Ausf√§lle gesamt", gesamt)
    c3.metric("√ò Ausf√§lle/Tag", f"{avg_tag:.2f}")
    st.markdown("---")

    # --- 2) Ausf√§lle nach Ausfallgrund ---
    st.subheader("üìä Ausf√§lle nach Grund")
    gr_counts = (
        df_filt["Ausfallgrund"]
        .value_counts()
        .reset_index(name="Anzahl")
        .rename(columns={"index": "Ausfallgrund"})
    )
    fig1 = px.bar(
        gr_counts,
        x="Ausfallgrund", y="Anzahl",
        title="Ausf√§lle nach Ausfallgrund",
        color="Ausfallgrund",
        color_discrete_sequence=px.colors.qualitative.Plotly
    )
    fig1.update_layout(xaxis_tickangle=45)
    st.plotly_chart(fig1, use_container_width=True)

    # --- 3) Ausf√§lle nach Bus ---
    st.subheader("üöç Ausf√§lle nach Bus")
    bus_counts = (
        df_filt["BusNr"]
        .value_counts()
        .reset_index(name="Anzahl")
        .rename(columns={"index": "BusNr"})
    )
    fig2 = px.bar(
        bus_counts,
        x="BusNr", y="Anzahl",
        title="Ausf√§lle pro Bus",
        color="Anzahl",
        color_continuous_scale="Viridis"
    )
    st.plotly_chart(fig2, use_container_width=True)

    # --- 4) Top-N Ausfallgr√ºnde als Pie ---
    st.subheader("ü•ß Top-Ausfallgr√ºnde")
    top_n = st.slider("Anzahl Top-Gr√ºnde", 3, 15, 7)
    top   = gr_counts.head(top_n)
    fig3 = px.pie(
        top,
        names="Ausfallgrund", values="Anzahl",
        title=f"Top {top_n} Ausfallgr√ºnde",
        color_discrete_sequence=px.colors.qualitative.Plotly
    )
    st.plotly_chart(fig3, use_container_width=True)
    st.markdown("---")

    # --- 5) Heatmap Status ---
    plot_status_heatmap(df_filt)
    df_dates = _load_registration_dates()  # oder Dein DataFrame mit Zulassung/Verkauf
    plot_series_status_heatmap(
    df_filt=df_filt,
    df_dates=df_dates,
    dpi=200,
    figsize=(16.53, 11.69),
    filename="heatmap_A3.png"
)
    st.markdown("---")

    # --- 6) Fehlerkategorien ---
    st.subheader("üè∑Ô∏è Auswertung nach Fehlerkategorien")
    df_cat = df_filt.copy()
    df_cat["Fehlerkategorie"] = (
        df_cat["Ausfallgrund"]
        .replace(FEHLERKATEGORIEN)
        .fillna("Sonstiges")
    )
    cat_counts = (
        df_cat["Fehlerkategorie"]
        .value_counts()
        .reset_index(name="Anzahl")
        .rename(columns={"index": "Fehlerkategorie"})
    )
    fig4 = px.bar(
        cat_counts,
        x="Fehlerkategorie", y="Anzahl",
        title="Anzahl der Ausf√§lle nach Fehlerkategorie",
        color="Fehlerkategorie",
        color_discrete_sequence=px.colors.qualitative.Plotly
    )
    fig4.update_layout(xaxis_tickangle=45)
    st.plotly_chart(fig4, use_container_width=True)
    export_full_reports(
    df_filt=df_filt,
    bus_to_hersteller=BUS_TO_HERSTELLER,

)

def to_excel_bytes(df: pd.DataFrame) -> bytes:
    """
    Schreibt das DataFrame in einen Excel-Stream und gibt die Bytes zur√ºck.
    """
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Auswertung")
    
    return buffer.getvalue()




def export_full_reports(
    df_filt: pd.DataFrame,
    bus_to_hersteller: dict,
    default_availability: float = 1.0,
    date_path = Path(__file__).parent/"Zulassung-Verkauf.xlsx",
    sheet_dates: int = 0
) -> None:
    import pandas as pd
    import streamlit as st
    from io import BytesIO
    from openpyxl.styles import PatternFill
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, Reference

    # 1) Zulassungs-/Verkaufsdaten einlesen
    df_dates = pd.read_excel(date_path, sheet_name=sheet_dates, engine="openpyxl")
    df_dates = df_dates.rename(columns={
        "KOM-Nr.":      "BusNr",
        "Einsatz":      "ZulassungDatum",
        "Verkauf":      "VerkaufDatum"
    })
    df_dates["BusNr"]          = df_dates["BusNr"].astype(str).str.strip()
    df_dates["ZulassungDatum"] = pd.to_datetime(df_dates["ZulassungDatum"], errors="coerce")
    if "VerkaufDatum" in df_dates.columns:
        df_dates["VerkaufDatum"] = pd.to_datetime(df_dates["VerkaufDatum"], errors="coerce")
    else:
        df_dates["VerkaufDatum"] = pd.NaT

    # 2) Arbeitskopie, Status und Perioden erg√§nzen
    df = df_filt.copy()
    df["BusNr"] = df["BusNr"].astype(str).str.strip()
    if "Status" not in df.columns:
        if "Ausfall-Typ" in df.columns:
            df["Status"] = df["Ausfall-Typ"].apply(
                lambda x: "Ausgefallen" if x != "Fahren" else "Fahren"
            )
        else:
            raise KeyError("Weder 'Status' noch 'Ausfall-Typ' gefunden.")
    df["Jahr"]    = df["Datum"].dt.year
    df["Monat"]   = df["Datum"].dt.month
    df["Quartal"] = df["Datum"].dt.to_period("Q")

    # 3) Vollmatrix Bus√óTag im gefilterten Gesamtzeitraum
    start = df["Datum"].min()
    end   = df["Datum"].max()
    all_dates = pd.date_range(start, end, freq="D")
    buses     = df["BusNr"].unique()
    full = (
        pd.MultiIndex
          .from_product([buses, all_dates], names=["BusNr","Datum"])
          .to_frame(index=False)
    )

    # 4) auf Zulassung/Verkauf einschr√§nken
    full = full.merge(
        df_dates[["BusNr","ZulassungDatum","VerkaufDatum"]],
        on="BusNr", how="left"
    )
    mask = (
        (full["Datum"] >= full["ZulassungDatum"]) &
        (
            full["VerkaufDatum"].isna() |
            (full["Datum"] <= full["VerkaufDatum"])
        )
    ).fillna(False)
    full = full.loc[mask, ["BusNr","Datum"]]

    # 5) Merge mit Status+Serie, fehlende auff√ºllen
    cols = df[["BusNr","Datum","Status","Serie"]].drop_duplicates()
    full = full.merge(cols, on=["BusNr","Datum"], how="left")
    full["Status"] = full["Status"].fillna("Fahren")
    full["Serie"]  = full["Serie"].fillna("Unbekannt")
    full["Jahr"]    = full["Datum"].dt.year
    full["Monat"]   = full["Datum"].dt.month
    full["Quartal"] = full["Datum"].dt.to_period("Q")

    # 6) Verf√ºgbarkeiten berechnen
    def calc_verf(df_sub, period_col, group_col, extras):
        return (
            df_sub
            .groupby(extras + [group_col, period_col])["Status"]
            .apply(lambda x: (x=="Fahren").sum() / len(x))
            .reset_index(name="Verf")
        )

    vb_mon  = calc_verf(full, "Monat",   "BusNr", ["Jahr"])
    vb_qua  = calc_verf(full, "Quartal", "BusNr", ["Jahr"])
    vb_jahr = calc_verf(full, "Jahr",    "BusNr", [])

    # 7) Serien-Verf√ºgbarkeit
    ser_map = full[["BusNr","Serie"]].drop_duplicates()

    def calc_ser(df_bus, period):
        tmp = df_bus.merge(ser_map, on="BusNr", how="left")
        if period == "Jahr":
            return tmp.groupby(["Jahr","Serie"])["Verf"].mean().reset_index()
        else:
            return tmp.groupby(["Jahr", period, "Serie"])["Verf"].mean().reset_index()

    vs_mon  = calc_ser(vb_mon,  "Monat")
    vs_qua  = calc_ser(vb_qua,  "Quartal")
    vs_jahr = calc_ser(vb_jahr, "Jahr")

    # 8) Excel in BytesIO schreiben
    buffer = BytesIO()

    def write_pivot(df_sheet, writer, sheet_name, idx, cols):
        """
        idx:  Liste von Spalten, die in den Zeilen-Index sollen (z.B. ['BusNr'])
        cols: Liste von Spalten, die in die Spaltenk√∂pfe sollen (z.B. ['Jahr','Monat'])
        """
        pt = df_sheet.pivot_table(
            index=idx,
            columns=cols,
            values="Verf",
            aggfunc="mean"
        )
        pt.to_excel(writer, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        max_r, max_c = ws.max_row, ws.max_column

        # Leere Zellen schwarz f√ºllen
        black = PatternFill("solid", fgColor="000000")
        for row in ws.iter_rows(min_row=2, min_col=2, max_row=max_r, max_col=max_c):
            for c in row:
                if c.value is None:
                    c.fill = black

        # Farbskala und Prozentformat
        for cidx in range(2, max_c+1):
            letter = get_column_letter(cidx)
            ws.conditional_formatting.add(
                f"{letter}2:{letter}{max_r}",
                ColorScaleRule(
                    start_type="num", start_value=0.0, start_color="FF0000",
                    mid_type=  "num", mid_value=  0.5, mid_color=  "FFFF00",
                    end_type=  "num", end_value=  1.0, end_color= "00FF00"
                )
            )
            for r in range(2, max_r+1):
                cell = ws[f"{letter}{r}"]
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.00%"

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Bus‚ÄêSicht (Jahre nebeneinander)
        write_pivot(vb_mon,  writer, "Bus_Monat",   ["BusNr"], ["Jahr","Monat"])
        write_pivot(vb_qua,  writer, "Bus_Quartal", ["BusNr"], ["Jahr","Quartal"])
        write_pivot(vb_jahr, writer, "Bus_Jahr",    ["BusNr"], ["Jahr"])

        # Serien‚ÄêSicht (Jahre nebeneinander)
        write_pivot(vs_mon,  writer, "Serie_Monat",   ["Serie"], ["Jahr","Monat"])
        write_pivot(vs_qua,  writer, "Serie_Quartal", ["Serie"], ["Jahr","Quartal"])
        write_pivot(vs_jahr, writer, "Serie_Jahr",    ["Serie"], ["Jahr"])

        # Statistik‚ÄêSheet (√ò-Verf√ºgbarkeit aller Busse)
        stat = vb_mon.pivot_table(
            index="BusNr",
            columns="Monat",
            values="Verf",
            fill_value=0
        ).sort_index()
        stat["Min"] = stat.min(axis=1)
        stat["Max"] = stat.max(axis=1)
        stat["√ò"]   = stat.mean(axis=1)
        stat.to_excel(writer, sheet_name="Statistiken")
        ws_s = writer.sheets["Statistiken"]
        max_r, max_c = ws_s.max_row, ws_s.max_column

        # Farbskala f√ºr Statistik
        for cidx in range(2, max_c+1):
            letter = get_column_letter(cidx)
            ws_s.conditional_formatting.add(
                f"{letter}2:{letter}{max_r}",
                ColorScaleRule(
                    start_type="num", start_value=0.0, start_color="FFCCCC",
                    end_type=  "num", end_value= 1.0,   end_color="00CC00"
                )
            )
            for r in range(2, max_r+1):
                cell = ws_s[f"{letter}{r}"]
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.00%"

        # √ò-Zeile drunter und Chart
        avg_row = max_r + 2
        ws_s.cell(row=avg_row, column=1, value="√ò alle Busse")
        for cidx in range(2, max_c+1):
            letter = get_column_letter(cidx)
            ws_s.cell(row=avg_row, column=cidx, value=f"=AVERAGE({letter}2:{letter}{max_r})")
            ws_s[f"{letter}{avg_row}"].number_format = "0.00%"

        chart = LineChart()
        chart.title = "√ò Verf√ºgbarkeit aller Busse"
        chart.y_axis.number_format = "0.00%"
        cats = Reference(ws_s, min_col=2, max_col=max_c, min_row=1)
        data = Reference(ws_s, min_col=2, max_col=max_c, min_row=avg_row, max_row=avg_row)
        chart.set_categories(cats)
        chart.add_data(data, from_rows=True)
        ws_s.add_chart(chart, "O3")

    # 9) Download-Button anzeigen
    st.download_button(
        label="üì• Excel-Export Verf√ºgbarkeit & Statistiken",
        data=buffer.getvalue(),
        file_name="verfuegbarkeit_und_statistiken.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
def page_monatliche_auswertungen(
    df_unused: pd.DataFrame,
    bus_to_hersteller: dict,
    filter_file: str = "exportierte_auswertungen/ausfallgrund_filter.json"
) -> None:
    """
    Erzeugt f√ºr einen ausgew√§hlten Monat eine Excel-Datei mit
    Verf√ºgbarkeits‚Äê und Statistik‚ÄêSheets (Farben, %‚ÄêFormat, Chart),
    genau wie export_full_reports es tut.
    """

    st.title("üìÖ Monatliche Auswertungen (Verf√ºgbarkeit & Statistiken)")

    # --- 1) Daten neu laden (unabh√§ngig von globalen Filtern) ---
    df_all = get_data("Zusammenfassung.xlsx")

    # --- 2) Persistenter Ausfallgrund‚ÄêFilter (optional) ---
    Path(Path(filter_file).parent).mkdir(parents=True, exist_ok=True)
    alle_gruende = sorted(df_all["Ausfallgrund"].dropna().unique())

    # a) Default aus JSON lesen
    if Path(filter_file).exists():
        try:
            saved = json.load(open(filter_file, "r"))
            default = saved.get("gruende", alle_gruende)
        except:
            default = alle_gruende
    else:
        default = alle_gruende

    # b) Sidebar‚ÄêMultiselect
    selected = st.sidebar.multiselect(
        "Filter Ausfallgrund",
        options=alle_gruende,
        default=default
    )
    # c) Speichern
    try:
        json.dump({"gruende": selected}, open(filter_file, "w"))
    except:
        pass

    # --- 3) Auf die gew√§hlten Ausfallgr√ºnde einschr√§nken ---
    if selected:
        df = df_all[df_all["Ausfallgrund"].isin(selected)]
    else:
        df = df_all.copy()

    if df.empty:
        st.warning("Keine Daten f√ºr die gew√§hlten Ausfallgr√ºnde.")
        return

    # --- 4) Monat‚ÄêAuswahl (Jahr‚ÄêMonat) ---
    # Wir bauen eine Perioden‚ÄêSpalte und nehmen die eindeutigen Perioden
    df["JahrMonat"] = df["Datum"].dt.to_period("M")
    perioden = sorted(df["JahrMonat"].unique())
    sel_period = st.selectbox(
        "Monat ausw√§hlen",
        options=perioden,
        format_func=lambda p: f"{p.year}-{p.month:02d}"
    )

    # Slice f√ºr diesen Monat
    df_monat = df[df["JahrMonat"] == sel_period].copy()

    if df_monat.empty:
        st.info(f"Keine Daten im Monat {sel_period}.")
        return
    
    # Status‚ÄêSpalte erg√§nzen, falls nicht vorhanden
    if "Status" not in df_monat.columns:
        df_monat["Status"] = df_monat["Ausfall-Typ"] \
            .apply(lambda x: "Ausgefallen" if x != "Fahren" else "Fahren")
    
    st.markdown(f"### Auswertung f√ºr **{sel_period.year}-{sel_period.month:02d}**")
    
    # Jetzt klappt der Export ohne KeyError
    export_full_reports(
        df_filt=df_monat,
        bus_to_hersteller=bus_to_hersteller
    )
    
    
